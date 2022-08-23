import requests
import locale
import functools
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import NamedStyle, Font, Alignment
from pyuca import Collator

# Make GET request to get all countries in a JSON format
data = requests.get('https://restcountries.com/v3.1/all')
countries = data.json()

# Sort the countries alphabetically
# Use pyuca library to sort special characters
c = Collator()
def sort_key(item):
  return c.sort_key(str(item['name']['common']))

countries = sorted(countries, key=sort_key)

# Create the spreadsheet
wb = Workbook()
ws = wb.active
ws.title = "Countries"

# Style for the spreadsheet title
title = ws['A1']
title.font = Font(name='Arial', size=16,  bold=True, color='4F4F4F')
title.alignment = Alignment(horizontal="center", vertical="center")
title.value = "Countries List"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

# Create and and apply a style for column name
countries_rows_style = NamedStyle(name="country_row")
countries_rows_style.font = Font(color='808080', bold=True, size=12)

ws.cell(2,1, "Name").style = countries_rows_style
ws.cell(2,2, "Capital").style = countries_rows_style
ws.cell(2,3, "Area").style = countries_rows_style
ws.cell(2,4, "Currencies").style = countries_rows_style
ws.column_dimensions['A'].bestFit = True
ws.column_dimensions['B'].bestFit = True
ws.column_dimensions['C'].bestFit = True
ws.column_dimensions['D'].bestFit = True

# Add countries to the spreadsheet
# Use enumerate to get the country and it's index
for r, country in enumerate(countries):
	if country.get('name'):
		ws.cell(row=r+3, column=1, value=country['name']["common"])
	else:
		ws.cell(row=r+3, column=1, value='-')
	
	if country.get('capital'):
		ws.cell(row=r+3, column=2, value=country['capital'][0])
	else:
		ws.cell(row=r+3, column=2, value='-')

	if country.get('area'):
		# Format area number to use comma for decimal places and point for thousand separator.
		locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
		area = locale.currency(country['area'], grouping=True, symbol=False)
		ws.cell(row=r+3, column=3, value=area).alignment = Alignment(horizontal="right")
	else:
		ws.cell(row=r+3, column=3, value='-')
	
	if country.get('currencies'):
		# Get all currencies and join in one string separating by comma 
		currencies = ",".join([currencie for currencie in country["currencies"].keys()])
		ws.cell(row=r+3, column=4, value=currencies)
	else:
		ws.cell(row=r+3, column=4, value='-')

# Save spreadsheet to main folder
wb.save("countries.xlsx")
print('Spreadsheet saved to main folder!')